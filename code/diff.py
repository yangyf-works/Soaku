from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from patiencediff import PatienceSequenceMatcher
from rapidfuzz import fuzz
import hashlib

def stable_hash(row):
    return hashlib.md5(repr(row).encode()).hexdigest()

def row_similarity(a, b, threshold=0.5):
    cutoff = threshold * 100
    cnt = 0
    for x, y in zip(a, b):
        if fuzz.ratio(str(x), str(y), score_cutoff=cutoff):
            cnt += 1
    return cnt / max(len(a), 1)

def match_rows(sub_a, sub_b, threshold=0.5):
    """
    行同士を類似度でマッチングする
    戻り値:
        matched: [(i, j or None)]
        unmatched_b: [j, ...]
    """
    matched = []
    used_b = set()

    for i, row_a in enumerate(sub_a):
        best_j = None
        best_score = 0

        for j, row_b in enumerate(sub_b):
            if j in used_b:
                continue

            score = row_similarity(row_a, row_b, threshold)
            if score > best_score:
                best_score = score
                best_j = j

        if best_score >= threshold:
            matched.append((i, best_j))
            used_b.add(best_j)
        else:
            matched.append((i, None))

    unmatched_b = [j for j in range(len(sub_b)) if j not in used_b]

    return matched, unmatched_b

def compare_sheets(wb1, wb2):
    sheets1 = wb1.sheetnames  # 順序あり
    sheets2 = wb2.sheetnames  # 順序あり

    set1 = set(sheets1)
    set2 = set(sheets2)

    # 順序を保ったまま抽出
    added = [s for s in sheets2 if s not in set1]     # wb2順
    deleted = [s for s in sheets1 if s not in set2]   # wb1順
    common = [s for s in sheets2 if s in set1]        # wb2順基準

    return {
        "added": added,
        "deleted": deleted,
        "common": common
    }

def get_header_map(ws, headrow=1):
    """
    ヘッダ行を取得して
    {ヘッダ名: 列番号} の辞書を返す
    ※同名ヘッダは A, A_2, A_3... のようにする
    """
    header_map = {}
    name_count = {}  # 出現回数を管理

    row = next(ws.iter_rows(min_row=headrow, max_row=headrow, values_only=True))

    for col_idx, v in enumerate(row, start=1):
        if v is None or (isinstance(v, str) and not v.strip()):
            continue

        base_name = str(v)

        # 出現回数カウント
        if base_name not in name_count:
            name_count[base_name] = 1
            name = base_name
        else:
            name_count[base_name] += 1
            name = f"{base_name}_{name_count[base_name]}"

        header_map[name] = col_idx

    return header_map

def compare_columns(ws1, ws2, headrow=1):

    header1 = get_header_map(ws1, headrow)
    header2 = get_header_map(ws2, headrow)

    set1 = set(header1.keys())
    set2 = set(header2.keys())

    added = [header2[col] for col in (set2 - set1)]
    deleted = [header1[col] for col in (set1 - set2)]
    common_keys = set1 & set2

    common_cols_a = [header1[col] for col in common_keys]
    common_cols_b = [header2[col] for col in common_keys]
    pairs = list(zip(common_cols_a, common_cols_b))

    pairs.sort(key=lambda x: x[1])  # common_cols_bでソート
    if pairs:
        common_cols_a, common_cols_b = zip(*pairs)
    else:
        common_cols_a, common_cols_b = [], []

    return {
        "added": added,   
        "deleted": deleted,
        "common": [common_cols_a, common_cols_b]
    }

def normalize_row(row, common_cols):
    result = []
    append = result.append

    for col in common_cols:
        v = row[col-1]

        if v is None:
            append("")
        else:
            append(str(v))

    return tuple(result)

def find_last_row(ws, start_row, cols):
    for row_idx in range(ws.max_row, start_row - 1, -1):
        row = ws[row_idx]
        if any(row[col-1].value not in (None, "") for col in cols):
            return row_idx
    return start_row

def to_excel_row(idx, base):
    return idx + base

def to_excel_cell(col_idx, row_idx):
    return f"{fast_col_letter(col_idx)}{row_idx}"

col_cache = {}

def fast_col_letter(col):
    if col not in col_cache:
        col_cache[col] = get_column_letter(col)
    return col_cache[col]

def exceldiff(
         path1 , 
         path2 ,
         dataonly=False,
         headrow=1,
         threshold=0.5):
    
    base = headrow + 1
    wb1 = load_workbook(path1, data_only=dataonly, read_only=True)
    wb2 = load_workbook(path2, data_only=dataonly, read_only=True)

    result = compare_sheets(wb1, wb2)

    diff_result = {}
    if result["added"]:
        diff_result["sheet_added"] = result["added"]
    if result["deleted"]:
        diff_result["sheet_deleted"] = result["deleted"]

    sheet_modified= {}
    for sheet_name in result["common"]:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        colresult = compare_columns(ws1, ws2, headrow)

        if colresult["added"]:
            sheet_modified.setdefault(sheet_name, {}).setdefault("columns", {})["added"] = [
                    get_column_letter(col)
                    for col in sorted(colresult["added"])
                ]
            
        if colresult["deleted"]:
            sheet_modified.setdefault(sheet_name, {}).setdefault("columns", {})["deleted"] = [
                    get_column_letter(col)
                    for col in sorted(colresult["deleted"])
                ]
            
        common_cols_a=[]
        common_cols_b=[]
        if colresult["common"][0]:
            common_cols_a, common_cols_b = colresult["common"]
            last_a = find_last_row(ws1, headrow+1, common_cols_a)
            last_b = find_last_row(ws2, headrow+1, common_cols_b)
            rows_a = [
                normalize_row(r, common_cols_a)
                for r in ws1.iter_rows(min_row=headrow+1, max_row=last_a, values_only=True)
            ]
            rows_b = [
                normalize_row(r, common_cols_b)
                for r in ws2.iter_rows(min_row=headrow+1, max_row=last_b, values_only=True)
            ]
            rows_a_sha = [stable_hash(r) for r in rows_a]
            rows_b_sha = [stable_hash(r) for r in rows_b]

            sm = PatienceSequenceMatcher(None, rows_a_sha, rows_b_sha)
            added_rows = []
            removed_rows = []
            cellsdiff = []
            for tag, i1, i2, j1, j2 in sm.get_opcodes():
                if tag == "replace":
                    sub_a = rows_a[i1:i2]
                    sub_b = rows_b[j1:j2]

                    """
                    replaceブロックを解析して
                    行変更・行追加・行削除・セル変更を再構築する
                    """
                    matched, unmatched_b = match_rows(sub_a, sub_b, threshold)
                    
                    # 行削除・行変更
                    for i, j in matched:
                        abs_i = i1 + i
                        excel_i = to_excel_row(abs_i, base)
                        if j is None:
                            removed_rows.append(excel_i)
                        else:
                            abs_j = j1 + j
                            excel_j = to_excel_row(abs_j, base)

                            row_a = sub_a[i]
                            row_b = sub_b[j]

                            if row_a == row_b:
                                continue

                            for col_idx, (a, b) in enumerate(zip(row_a, row_b)):
                                if a != b:
                                    cellsdiff.append((
                                        to_excel_cell(common_cols_a[col_idx], excel_i),
                                        to_excel_cell(common_cols_b[col_idx], excel_j)
                                    ))

                    # 行追加
                    for j in unmatched_b:
                        abs_j = j1 + j
                        excel_j = to_excel_row(abs_j, base)
                        added_rows.append(excel_j)

                elif tag == "delete":
                    removed_rows.extend(to_excel_row(i, base) for i in range(i1, i2))

                elif tag == "insert":
                    added_rows.extend(to_excel_row(j, base) for j in range(j1, j2))

            if added_rows:
                sheet_modified.setdefault(sheet_name, {}).setdefault("rows", {})["added"] = sorted(added_rows)

            if removed_rows:
                sheet_modified.setdefault(sheet_name, {}).setdefault("rows", {})["deleted"] = sorted(removed_rows)
                
            if cellsdiff:
                sheet_modified.setdefault(sheet_name, {}).setdefault("rows", {}).update({"modified": cellsdiff})
            
    if sheet_modified:
        diff_result.setdefault("sheet_modified", {}).update(sheet_modified)
    return diff_result
